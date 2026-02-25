from typing import List, Tuple, Dict, Any, Optional
from io import BytesIO
import csv
import uuid
import aiomysql
from app.services.db_pool import db_pool
from app.controllers.conn import conn_controller
from app.settings.config import settings

# SIM导入固定连接的Id
sim_conn_id = settings.SIM_CONN_ID

_PROGRESS: Dict[str, Dict[str, Any]] = {}

class SIMService:
    """SIM-ICCID导入服务"""

    def _progress_start(self, stamp: str, filename: str):
        """开始任务"""
        _PROGRESS[stamp] = {
            "file": filename,
            "stage": "parsing",
            "total": 0,
            "current": 0,
            "message": "开始解析文件",
            "success": False,
        }

    def _progress_update(self, stamp: str, **kwargs):
        """更新进度"""
        if stamp in _PROGRESS:
            _PROGRESS[stamp].update(kwargs)

    def _progress_fail(self, stamp: str, message: str):
        """失败"""
        self._progress_update(stamp, stage="failed", message=message, success=False)

    def _progress_done(self, stamp: str):
        """完成"""
        self._progress_update(stamp, stage="done", message="处理完成", success=True)

    def get_progress(self, stamp: str) -> Dict[str, Any]:
        """获取进度"""
        return _PROGRESS.get(stamp, {"stage": "", "message": ""})

    async def _ensure_pool(self) -> None:
        """确保连接池已注册"""
        pool = db_pool.get_pool(sim_conn_id)
        if pool is not None:
            try:
                if hasattr(pool, "closed") and pool.closed:
                    pool = None
                elif hasattr(pool, "_closed") and pool._closed:
                    pool = None
            except Exception:
                pool = None
        if pool is not None:
            return
        conn = await conn_controller.get_decrypted_connection(sim_conn_id)
        if not conn:
            raise ValueError("连接记录不存在或密码无法解密，请在连接管理重新保存密码")
        await db_pool.register_pool(
            conn_id=conn["id"],
            db_type=conn["db_type"],
            host=conn["host"],
            port=conn["port"],
            username=conn["username"],
            password=conn["password"],
            database=conn["database"],
            params=conn["params"],
        )

    async def _insert_rows(self, stamp: str, rows: List[Tuple[str, str]]) -> int:
        """批量写入临时表"""
        if not rows:
            return 0
        await self._ensure_pool()
        self._progress_update(stamp, stage="writing", total=len(rows), current=0, message="写入临时表")
        pool = db_pool.get_pool(sim_conn_id)
        if pool is None:
            raise ValueError("连接池不存在")
        if isinstance(pool, aiomysql.Pool):
            async with pool.acquire() as conn:
                async with conn.cursor() as cur:
                    await cur.executemany(
                        "INSERT INTO tm_simiccidimp (ImpStamp, SimNumber, ICCID) VALUES (%s, %s, %s)",
                        [(stamp, r[0], r[1]) for r in rows],
                    )
                    await conn.commit()
                    self._progress_update(stamp, current=len(rows))
                    return cur.rowcount or 0
        raise ValueError("不支持的连接池类型")

    async def upload_excel(self, file_bytes: bytes, filename: str, stamp: Optional[str] = None) -> str:
        """上传Excel并写入临时表"""
        if stamp is None:
            stamp = str(uuid.uuid4())
        self._progress_start(stamp, filename)
        ext = filename.split(".")[-1].lower() if "." in filename else ""
        rows: List[Tuple[str, str]] = []
        if ext in ("xlsx",):
            try:
                import openpyxl  # type: ignore
            except Exception:
                raise ValueError("服务器未安装Excel解析库，请上传CSV或联系管理员安装openpyxl")
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
            if headers != ["SimNumber", "ICCID"]:
                raise ValueError("Excel列名不符合要求，需为SimNumber、ICCID")
            self._progress_update(stamp, stage="parsing", message="解析Excel中")
            for row in ws.iter_rows(min_row=2):
                sim = str(row[0].value).strip() if row[0].value is not None else ""
                iccid = str(row[1].value).strip() if row[1].value is not None else ""
                if sim or iccid:
                    rows.append((sim, iccid))
        elif ext in ("csv",):
            text = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.reader(BytesIO(text.encode("utf-8")))
            try:
                headers = next(reader)
            except StopIteration:
                headers = []
            headers = [h.strip() for h in headers]
            if headers != ["SimNumber", "ICCID"]:
                raise ValueError("CSV列名不符合要求，需为SimNumber、ICCID")
            self._progress_update(stamp, stage="parsing", message="解析CSV中")
            for r in reader:
                if not r or len(r) < 2:
                    continue
                sim = (r[0] or "").strip()
                iccid = (r[1] or "").strip()
                if sim or iccid:
                    rows.append((sim, iccid))
        else:
            raise ValueError("仅支持.xlsx或.csv文件")
        await self._insert_rows(stamp, rows)
        return stamp

    async def process_tmp(self, stamp: str) -> None:
        """调用存储过程处理临时表数据"""
        await self._ensure_pool()
        pool = db_pool.get_pool(sim_conn_id)
        if pool is None:
            raise ValueError("连接池不存在")
        if isinstance(pool, aiomysql.Pool):
            async with pool.acquire() as conn:
                async with conn.cursor() as cur:
                    try:
                        self._progress_update(stamp, stage="processing", message="调用存储过程")
                        await cur.execute("CALL proc_ImportSimICCID(%s)", (stamp,))
                        result_row = await cur.fetchone()
                        er_type = 0
                        er_message = ""
                        if result_row is not None and len(result_row) >= 1:
                            try:
                                er_type = int(result_row[0])
                            except Exception:
                                er_type = 0
                            if len(result_row) > 1 and result_row[1]:
                                er_message = str(result_row[1])
                        while True:
                            has_next = await cur.nextset()
                            if not has_next:
                                break
                        await conn.commit()
                        if er_type != 0:
                            self._progress_fail(stamp, er_message or "存储过程执行失败")
                            raise ValueError(er_message or "存储过程执行失败")
                        if er_message:
                            self._progress_update(stamp, message=er_message)
                        self._progress_done(stamp)
                    except Exception as e:
                        await conn.commit()
                        self._progress_fail(stamp, str(e))
                        raise
            return
        raise ValueError("不支持的连接池类型")

    async def submit_and_process(self, file_bytes: bytes, filename: str, stamp: Optional[str] = None) -> str:
        """后台执行：上传并处理"""
        if stamp is None:
            stamp = str(uuid.uuid4())
        try:
            await self.upload_excel(file_bytes, filename, stamp=stamp)
            await self.process_tmp(stamp)
        except Exception as e:
            self._progress_fail(stamp, str(e))
            raise
        return stamp


sim_service = SIMService()
