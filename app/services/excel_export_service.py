"""
Excel导出服务 - 支持大数据量导出、分sheet、分文件、ZIP压缩
"""
import os
import zipfile
import asyncio
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.models.report import ReportGeneration, ReportConfig
from app.models.conn import DBConnection
from app.services.sql_execution_service import SQLExecutionService
from app.services.oss_service import oss_service
from app.log import logger
from app.core.config_loader import config
from app.settings import settings


class ExcelExportService:
    """Excel导出服务"""

    # 从配置读取常量
    MAX_ROWS_PER_SHEET = config.report.max_rows_per_sheet
    MAX_SHEETS_PER_FILE = config.report.max_sheets_per_file
    PAGE_SIZE = config.report.page_size
    STREAM_MAX_ROWS_PER_SHEET = config.report.stream_max_rows_per_sheet
    STREAM_MAX_SHEETS_PER_FILE = config.report.stream_max_sheets_per_file
    STREAM_FULL_STYLE_MAX_ROWS = config.report.stream_full_style_max_rows
    ZIP_THRESHOLD_BYTES = 10 * 1024 * 1024  # 单文件超过10MB则压缩

    async def export_report(self, generation_id: int):
        """
        导出报表主流程
        :param generation_id: 报表生成记录ID
        """
        generation = None
        try:
            logger.info(f"开始导出报表, generation_id: {generation_id}")
            
            # 获取生成记录
            generation = await ReportGeneration.get_or_none(id=generation_id).prefetch_related("report_config")
            if not generation:
                logger.error(f"报表生成记录不存在: {generation_id}")
                return

            logger.info(f"获取到生成记录: {generation.report_name}")
            generation.status = "exporting"
            generation.progress = 1
            generation.progress_text = "任务启动"
            generation.exported_rows = 0
            generation.error_message = None
            await generation.save(update_fields=["status", "progress", "progress_text", "exported_rows", "error_message"])

            # 获取报表配置
            config = await generation.report_config
            logger.info(f"获取报表配置: {config}")
            if not config:
                logger.error("报表配置不存在")
                await self._update_generation_status(
                    generation,
                    "failed",
                    error_msg="报表配置不存在"
                )
                return

            logger.info(f"报表配置: {config.report_name}")

            # 获取数据库连接
            db_conn = await config.db_connection
            logger.info(f"获取数据库连接: {db_conn}")
            if not db_conn:
                logger.error("数据库连接不存在")
                await self._update_generation_status(
                    generation,
                    "failed",
                    error_msg="数据库连接不存在"
                )
                return

            logger.info(f"数据库连接: {db_conn.name}")

            # 记录执行日志
            execution_log = {
                "sql": config.sql_statement,
                "db_connection": {
                    "id": db_conn.id,
                    "name": db_conn.name,
                    "type": db_conn.db_type,
                    "host": db_conn.host,
                    "database": db_conn.database
                },
                "start_time": datetime.now().isoformat()
            }

            # 执行导出
            local_file_path = await self._execute_export(
                generation=generation,
                db_conn=db_conn,
                sql=config.sql_statement
            )
            file_path = local_file_path

            # 上传到OSS（启用时）
            oss_meta = await asyncio.to_thread(oss_service.upload_file, local_file_path)
            if oss_meta:
                execution_log["storage"] = oss_meta
                execution_log["local_file_path"] = local_file_path
                execution_log["storage_upload_status"] = "success"
                # 仍保留本地file_path用于兼容历史逻辑，下载优先走OSS返回直链
                if settings.OSS_CLEANUP_LOCAL_AFTER_UPLOAD and os.path.exists(local_file_path):
                    try:
                        os.remove(local_file_path)
                        execution_log["local_file_cleaned"] = True
                    except Exception as cleanup_exc:
                        logger.warning(f"清理本地报表文件失败: {local_file_path}, error={cleanup_exc}")
            elif settings.OSS_ENABLED:
                execution_log["storage_upload_status"] = "failed"
                execution_log["storage_upload_error"] = "OSS上传失败，已回退本地文件地址"
                execution_log["local_file_path"] = local_file_path
                logger.error(
                    f"OSS上传失败，报表将保留本地下载路径: generation_id={generation_id}, file={local_file_path}"
                )

            # 更新执行日志
            execution_log["end_time"] = datetime.now().isoformat()
            execution_log["status"] = "success"
            if oss_meta:
                remote_path = (
                    oss_meta.get("download_url")
                    or oss_meta.get("url")
                    or oss_meta.get("oss_uri")
                )
                execution_log["file_path"] = remote_path or file_path
            else:
                execution_log["file_path"] = file_path

            # 更新生成记录状态
            generation.status = "completed"
            generation.completed_at = datetime.now()
            generation.progress = 100
            generation.progress_text = "导出完成"
            generation.error_message = None
            generation.file_path = file_path
            generation.execution_json = execution_log
            await generation.save()

            logger.info(f"报表导出成功: {generation.report_name}, 文件: {file_path}")

        except Exception as e:
            logger.error(f"报表导出失败: {str(e)}", exc_info=True)
            if generation:
                try:
                    await self._update_generation_status(
                        generation,
                        "failed",
                        error_msg=str(e)
                    )
                except Exception as update_error:
                    logger.error(f"更新状态失败: {str(update_error)}")

    async def _execute_export(
        self,
        generation: ReportGeneration,
        db_conn: DBConnection,
        sql: str
    ) -> str:
        """
        执行导出逻辑
        :return: 生成的文件路径
        """
        logger.info(f"_execute_export 开始, sql长度: {len(sql)}")
        if db_conn.db_type == "mysql":
            return await self._execute_export_mysql_stream(generation, db_conn, sql)
        
        logger.info("大报表导出使用无count分页模式，避免COUNT(*)拖慢导出")

        # 创建文件存储目录
        file_dir = self._get_file_dir()
        os.makedirs(file_dir, exist_ok=True)

        # 生成的文件列表
        file_list = []

        # 分批导出
        current_row = 0
        current_sheet = 0
        current_file = 0
        wb = None
        ws = None
        headers = None
        sheet_data = []  # 存储当前sheet的数据

        try:
            has_any_data = False
            done = False
            while not done:
                # 判断是否需要创建新文件
                if current_sheet % self.MAX_SHEETS_PER_FILE == 0:
                    # 保存上一个文件
                    if wb:
                        file_path = await asyncio.to_thread(
                            self._save_workbook,
                            wb,
                            file_dir,
                            generation.report_name,
                            current_file
                        )
                        file_list.append(file_path)

                    # 创建新工作簿（普通模式，支持样式）
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)  # 删除默认sheet
                    current_file += 1
                    logger.info(f"创建第 {current_file} 个Excel文件")

                # 创建新sheet
                ws = wb.create_sheet(title=f"Sheet{current_sheet + 1}")
                current_sheet += 1
                sheet_data = []  # 重置sheet数据
                logger.info(f"创建第 {current_sheet} 个sheet")

                # 收集数据到当前sheet
                rows_written, headers = await self._collect_data_to_sheet(
                    ws=ws,
                    db_conn=db_conn,
                    sql=sql,
                    offset=current_row,
                    limit=self.MAX_ROWS_PER_SHEET,
                    headers=headers
                )

                if rows_written == 0:
                    # 当前sheet没有数据，移除空sheet并结束
                    wb.remove(ws)
                    done = True
                    break

                has_any_data = True
                current_row += rows_written
                logger.info(f"Sheet {current_sheet} 写入 {rows_written} 行，累计 {current_row}")
                if current_row % self.PAGE_SIZE == 0:
                    await self._update_generation_progress(
                        generation=generation,
                        progress=min(95, 10 + current_row // 100000),
                        progress_text=f"导出中：已导出 {current_row} 行",
                        exported_rows=current_row,
                    )
                    # 长时间任务需要定期yield控制权，避免阻塞事件循环
                    await asyncio.sleep(0.01)

                # 当前sheet未写满，说明已到末尾
                if rows_written < self.MAX_ROWS_PER_SHEET:
                    done = True

            # 保存最后一个文件
            if wb and wb.worksheets:
                file_path = await asyncio.to_thread(
                    self._save_workbook,
                    wb,
                    file_dir,
                    generation.report_name,
                    current_file
                )
                file_list.append(file_path)

            if not has_any_data:
                raise ValueError("查询结果为空，无法导出")

            await self._update_generation_progress(
                generation=generation,
                progress=97,
                progress_text="压缩打包中",
                exported_rows=current_row,
            )
            final_path = await self._finalize_export_files(
                file_list=file_list,
                file_dir=file_dir,
                report_name=generation.report_name
            )
            await self._update_generation_progress(
                generation=generation,
                progress=99,
                progress_text="文件处理完成，等待收尾",
                exported_rows=current_row,
            )
            return final_path

        except Exception as e:
            # 清理临时文件
            for file_path in file_list:
                if os.path.exists(file_path):
                    os.remove(file_path)
            raise

    async def _execute_export_mysql_stream(
        self,
        generation: ReportGeneration,
        db_conn: DBConnection,
        sql: str,
    ) -> str:
        """
        MySQL流式导出：单次执行SQL，避免深分页OFFSET导致的大SQL卡死问题。
        """
        logger.info("MySQL报表导出使用流式拉取模式")
        file_dir = self._get_file_dir()
        os.makedirs(file_dir, exist_ok=True)
        file_list: List[str] = []
        wb = None
        ws = None
        current_sheet = 0
        current_file = 0
        headers = None
        original_headers = None
        exported_rows = 0
        rows_in_sheet = 0
        styled_sheets = set()
        stream_rows_per_sheet = max(1000, self.STREAM_MAX_ROWS_PER_SHEET)
        stream_sheets_per_file = max(1, self.STREAM_MAX_SHEETS_PER_FILE)

        async def _start_new_sheet():
            nonlocal wb, ws, current_sheet, current_file, rows_in_sheet
            if wb is None or current_sheet % stream_sheets_per_file == 0:
                if wb and wb.worksheets:
                    file_path = await asyncio.to_thread(
                        self._save_workbook,
                        wb,
                        file_dir,
                        generation.report_name,
                        current_file
                    )
                    file_list.append(file_path)
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                current_file += 1
                logger.info(f"创建第 {current_file} 个Excel文件（流式）")
            current_sheet += 1
            ws = wb.create_sheet(title=f"Sheet{current_sheet}")
            rows_in_sheet = 0
            logger.info(f"创建第 {current_sheet} 个sheet（流式）")

        try:
            async for batch in SQLExecutionService.execute_query_stream_mysql(
                db_conn=db_conn,
                sql=sql,
                batch_size=self.PAGE_SIZE,
            ):
                if not batch:
                    continue
                if headers is None:
                    original_headers = list(batch[0].keys())
                    headers = self._build_unique_headers(original_headers)
                for row in batch:
                    if ws is None:
                        await _start_new_sheet()
                        if headers:
                            for col_idx, header in enumerate(headers, 1):
                                ws.cell(row=1, column=col_idx, value=header)
                    ws.append([row.get(h) for h in original_headers])
                    rows_in_sheet += 1
                    exported_rows += 1
                    if exported_rows % self.PAGE_SIZE == 0:
                        await self._update_generation_progress(
                            generation=generation,
                            progress=min(95, 10 + exported_rows // 100000),
                            progress_text=f"导出中：已导出 {exported_rows} 行",
                            exported_rows=exported_rows,
                        )
                        await asyncio.sleep(0)
                    if rows_in_sheet >= stream_rows_per_sheet:
                        await self._update_generation_progress(
                            generation=generation,
                            progress=min(95, 15 + exported_rows // 100000),
                            progress_text=f"导出中：已完成第 {current_sheet} 个Sheet，累计 {exported_rows} 行",
                            exported_rows=exported_rows,
                        )
                        if ws and current_sheet not in styled_sheets:
                            if rows_in_sheet > self.STREAM_FULL_STYLE_MAX_ROWS:
                                await asyncio.to_thread(self._apply_sheet_style_light, ws, headers)
                            else:
                                await asyncio.to_thread(self._apply_sheet_style, ws, headers)
                            styled_sheets.add(current_sheet)
                        ws = None
                        rows_in_sheet = 0

            if exported_rows == 0:
                raise ValueError("查询结果为空，无法导出")

            if ws and current_sheet not in styled_sheets:
                if rows_in_sheet > self.STREAM_FULL_STYLE_MAX_ROWS:
                    await asyncio.to_thread(self._apply_sheet_style_light, ws, headers)
                else:
                    await asyncio.to_thread(self._apply_sheet_style, ws, headers)
                styled_sheets.add(current_sheet)

            if wb and wb.worksheets:
                file_path = await asyncio.to_thread(
                    self._save_workbook,
                    wb,
                    file_dir,
                    generation.report_name,
                    current_file
                )
                file_list.append(file_path)

            await self._update_generation_progress(
                generation=generation,
                progress=97,
                progress_text="压缩打包中",
                exported_rows=exported_rows,
            )
            final_path = await self._finalize_export_files(
                file_list=file_list,
                file_dir=file_dir,
                report_name=generation.report_name
            )
            await self._update_generation_progress(
                generation=generation,
                progress=99,
                progress_text="文件处理完成，等待收尾",
                exported_rows=exported_rows,
            )
            return final_path
        except Exception:
            for file_path in file_list:
                if os.path.exists(file_path):
                    os.remove(file_path)
            raise

    async def _collect_data_to_sheet(
        self,
        ws,
        db_conn: DBConnection,
        sql: str,
        offset: int,
        limit: int,
        headers: Optional[List[str]] = None
    ) -> tuple[int, List[str]]:
        """
        收集数据并写入sheet，应用样式美化
        :return: (写入的行数, 表头列表)
        """
        rows_written = 0
        batch_size = self.PAGE_SIZE
        original_headers = None  # 原始字段名，用于数据获取
        all_data = []  # 收集所有数据

        # 分批查询数据
        while rows_written < limit:
            current_batch_size = min(batch_size, limit - rows_written)

            data = await SQLExecutionService.execute_query_page(
                db_conn=db_conn,
                sql=sql,
                offset=offset + rows_written,
                limit=current_batch_size
            )

            if not data:
                break

            # 记录表头（仅第一次）
            if rows_written == 0 and headers is None:
                original_headers = list(data[0].keys())
                headers = self._build_unique_headers(original_headers)
            elif original_headers is None:
                original_headers = list(data[0].keys())

            all_data.extend(data)
            rows_written += len(data)
            if rows_written % self.PAGE_SIZE == 0:
                await asyncio.sleep(0)

            if len(data) < current_batch_size:
                break

        # 写入数据并应用样式
        if headers and all_data:
            await asyncio.to_thread(self._write_with_styles, ws, headers, all_data, original_headers)

        return rows_written, headers

    def _write_with_styles(
        self,
        ws,
        headers: List[str],
        data: List[Dict],
        original_headers: List[str]
    ):
        """
        写入数据并应用Excel样式美化
        """
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 交替行颜色
        odd_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # 浅蓝
        even_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色

        # 边框
        thin_border = Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据行
        for row_idx, row_data in enumerate(data, 2):
            fill = odd_fill if row_idx % 2 == 0 else even_fill
            for col_idx, h in enumerate(original_headers, 1):
                value = row_data.get(h)
                # None值显示为空字符串
                if value is None:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.fill = fill
                cell.alignment = data_alignment
                cell.border = thin_border

        self._apply_sheet_style(ws, headers)

    def _auto_adjust_column_width(
        self,
        ws,
        headers: List[str],
        data: List[Dict],
        original_headers: List[str]
    ):
        """
        自动调整列宽
        """
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            # 检查数据行，最多检查前100行
            for row_data in data[:100]:
                value = row_data.get(original_headers[col_idx - 1])
                if value:
                    # 中文字符算2个宽度
                    cell_length = 0
                    for char in str(value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 2
                        else:
                            cell_length += 1
                    max_length = max(max_length, cell_length)
            # 设置列宽，最小8，最大50
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    def _display_width(self, value: Any) -> int:
        text = "" if value is None else str(value)
        width = 0
        for char in text:
            if '\u4e00' <= char <= '\u9fff':
                width += 2
            else:
                width += 1
        return width

    def _apply_sheet_style(self, ws, headers: Optional[List[str]]):
        """
        对已有sheet应用统一样式：
        - 首行冻结
        - 自动筛选
        - 表头加粗+大一号
        - 行高列宽自适配（近似）
        - 套用表格样式
        """
        if not headers:
            return

        max_row = ws.max_row
        max_col = len(headers)
        if max_row <= 0 or max_col <= 0:
            return

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )

        # 表头样式
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据区域样式 + 列宽估算
        sample_limit = min(max_row, 2000)
        col_widths = [self._display_width(h) for h in headers]
        for row_idx in range(2, max_row + 1):
            row_max_lines = 1
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = data_alignment
                cell.border = thin_border
                if row_idx <= sample_limit:
                    v = "" if cell.value is None else str(cell.value)
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], self._display_width(v))
                    row_max_lines = max(row_max_lines, v.count("\n") + 1)
            if row_max_lines > 1:
                ws.row_dimensions[row_idx].height = min(120, 18 * row_max_lines)

        # 表头行高
        ws.row_dimensions[1].height = 24

        # 自动列宽（近似）
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 8), 60)

        # 冻结首行 + 全列筛选
        max_col_letter = get_column_letter(max_col)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

        # 表格样式（避免重复创建）
        table_name = f"Table_{ws.title}".replace(" ", "_").replace("-", "_")
        if table_name not in ws.tables:
            table = Table(displayName=table_name, ref=f"A1:{max_col_letter}{max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

    def _apply_sheet_style_light(self, ws, headers: Optional[List[str]]):
        """
        大数据量轻量样式：
        - 保留表头样式、首行冻结、筛选、基础列宽
        - 跳过全量单元格样式遍历和表格样式，降低内存和CPU压力
        """
        if not headers:
            return

        max_row = ws.max_row
        max_col = len(headers)
        if max_row <= 0 or max_col <= 0:
            return

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 24

        # 仅基于表头和少量样本估算列宽，避免全量遍历
        sample_limit = min(max_row, 300)
        col_widths = [self._display_width(h) for h in headers]
        for row_idx in range(2, sample_limit + 1):
            for col_idx in range(1, max_col + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], self._display_width(v))

        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 8), 50)

        max_col_letter = get_column_letter(max_col)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

    def _save_workbook(
        self,
        wb: Workbook,
        file_dir: str,
        report_name: str,
        file_index: int
    ) -> str:
        """
        保存工作簿
        """
        # 替换文件名中的非法字符
        safe_name = report_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        file_name = f"{safe_name}_{file_index}.xlsx"
        file_path = os.path.join(file_dir, file_name)
        wb.save(file_path)
        logger.info(f"保存Excel文件: {file_path}")
        return file_path

    def _build_unique_headers(self, original_headers: List[str]) -> List[str]:
        seen = {}
        unique_headers = []
        for h in original_headers:
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 1
                unique_headers.append(h)
        return unique_headers

    def _compress_if_large(self, file_path: str) -> str:
        """
        如果单个报表文件超过阈值，则压缩成zip并删除原文件。
        :return: 最终文件路径（xlsx或zip）
        """
        if not os.path.exists(file_path):
            return file_path
        file_size = os.path.getsize(file_path)
        if file_size <= self.ZIP_THRESHOLD_BYTES:
            return file_path

        base_name = os.path.splitext(os.path.basename(file_path))[0]
        zip_path = os.path.join(os.path.dirname(file_path), f"{base_name}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, os.path.basename(file_path))
        os.remove(file_path)
        logger.info(f"单文件超过10MB，已压缩: {zip_path}")
        return zip_path

    async def _finalize_export_files(self, file_list: List[str], file_dir: str, report_name: str) -> str:
        if len(file_list) > 1:
            safe_name = report_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            zip_path = os.path.join(file_dir, f"{safe_name}.zip")
            await asyncio.to_thread(self._zip_files, zip_path, file_list)
            await asyncio.to_thread(self._remove_files, file_list)
            logger.info(f"生成ZIP文件: {zip_path}")
            return zip_path
        return await asyncio.to_thread(self._compress_if_large, file_list[0])

    def _zip_files(self, zip_path: str, file_list: List[str]):
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in file_list:
                zipf.write(file_path, os.path.basename(file_path))

    def _remove_files(self, file_list: List[str]):
        for file_path in file_list:
            if os.path.exists(file_path):
                os.remove(file_path)

    async def _update_generation_progress(
        self,
        generation: ReportGeneration,
        progress: int,
        progress_text: str,
        exported_rows: int,
    ):
        try:
            generation.progress = max(0, min(progress, 99))
            generation.progress_text = progress_text
            generation.exported_rows = exported_rows
            await generation.save(update_fields=["progress", "progress_text", "exported_rows"])
        except Exception as e:
            logger.warning(f"更新导出进度失败: {str(e)}")

    def _get_file_dir(self) -> str:
        """
        获取文件存储目录
        """
        now = datetime.now()
        dir_path = os.path.join(
            config.report.report_dir,
            str(now.year),
            f"{now.month:02d}",
            f"{now.day:02d}"
        )
        return dir_path

    async def _update_generation_status(
        self,
        generation: ReportGeneration,
        status: str,
        error_msg: Optional[str] = None
    ):
        """
        更新生成记录状态
        """
        try:
            generation.status = status
            generation.completed_at = datetime.now()
            if status == "failed":
                short_error = (error_msg or "未知错误").strip()
                generation.progress_text = f"导出失败: {short_error[:120]}"
                generation.error_message = short_error
            else:
                generation.error_message = None

            if error_msg:
                execution_log = generation.execution_json or {}
                execution_log["error"] = error_msg
                execution_log["end_time"] = datetime.now().isoformat()
                generation.execution_json = execution_log

            await generation.save()
        except Exception as e:
            logger.error(f"更新生成记录状态失败: {str(e)}")
