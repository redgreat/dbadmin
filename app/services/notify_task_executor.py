import asyncio
import hashlib
import hmac
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl

from app.core.config_loader import config
from app.models.alert import AlertSendLog, AlertSender
from app.models.conn import DBConnection
from app.models.report import ReportConfig, ReportGeneration
from app.models.task_notify import NotifyTaskRunLog, ReportSendTask, SqlAlertTask, TaskRunStatus
from app.services.excel_export_service import ExcelExportService
from app.services.oss_service import oss_service
from app.services.report_service import ReportService
from app.services.sql_execution_service import SQLExecutionService
from app.services.wecom_bot_service import WecomBotService
from app.settings import settings


class NotifyTaskExecutor:
    MAX_MESSAGE_ROWS = 20
    _report_send_running_task_ids = set()
    _report_send_running_lock = asyncio.Lock()

    @classmethod
    async def _acquire_report_send_lock(cls, task_id: int) -> bool:
        async with cls._report_send_running_lock:
            if task_id in cls._report_send_running_task_ids:
                return False
            cls._report_send_running_task_ids.add(task_id)
            return True

    @classmethod
    async def _release_report_send_lock(cls, task_id: int):
        async with cls._report_send_running_lock:
            cls._report_send_running_task_ids.discard(task_id)

    @staticmethod
    async def _create_run_log(task_type: str, task_ref_id: int) -> NotifyTaskRunLog:
        return await NotifyTaskRunLog.create(
            task_type=task_type,
            task_ref_id=task_ref_id,
            status=TaskRunStatus.RUNNING,
            start_time=datetime.now(),
        )

    @staticmethod
    async def _finish_run_log(
        run_log: NotifyTaskRunLog,
        status: str,
        output: Optional[str] = None,
        error: Optional[str] = None,
        result_json: Optional[dict] = None,
    ):
        run_log.status = status
        run_log.output = output
        run_log.error = error
        run_log.result_json = result_json
        start_time = run_log.start_time
        end_time = datetime.now(tz=start_time.tzinfo) if getattr(start_time, "tzinfo", None) else datetime.now()
        run_log.end_time = end_time
        if start_time and end_time:
            safe_start = start_time
            safe_end = end_time
            if (getattr(safe_start, "tzinfo", None) is None) != (getattr(safe_end, "tzinfo", None) is None):
                if getattr(safe_start, "tzinfo", None) is None:
                    safe_start = safe_start.replace(tzinfo=safe_end.tzinfo)
                else:
                    safe_end = safe_end.replace(tzinfo=safe_start.tzinfo)
            run_log.duration = int((safe_end - safe_start).total_seconds())
        await run_log.save()

    @staticmethod
    async def _create_alert_send_log(
        sender: AlertSender,
        alert_text: str,
        send_status: int,
        response_text: Optional[str] = None,
    ):
        await AlertSendLog.create(
            sender_id=sender.id,
            sender_name=sender.sender_name,
            channel_type=sender.channel_type,
            channel_target=sender.channel_target,
            alert_text=alert_text,
            send_status=send_status,
            response_text=response_text,
            sent_at=datetime.now(),
        )

    @staticmethod
    def _fill_report_message(template: str, report_name: str) -> str:
        """渲染定时报表消息模板占位符。"""
        now = datetime.now()
        msg = (template or "").strip()
        if not msg:
            return (
                f"定时报表已生成\n"
                f"报表：{report_name}\n"
                f"生成时间：{now.strftime('%Y-%m-%d %H:%M:%S')}"
            )

        replacements = {
            "{{date}}": now.strftime("%Y%m%d"),
            "{{time}}": now.strftime("%H%M"),
            "{{month}}": now.strftime("%Y%m"),
            "{{report_name}}": report_name,
        }
        for key, value in replacements.items():
            msg = msg.replace(key, value)
        return msg

    @staticmethod
    def _guess_public_base_url() -> str:
        """推断用于外部访问的服务根地址。"""
        try:
            if getattr(config, "nginx", None) and config.nginx.enabled:
                host = (config.nginx.server_name or "localhost").strip()
                port = int(config.nginx.listen_port)
            else:
                host = (getattr(config, "server", None) and config.server.host) or "localhost"
                host = "localhost" if host in {"0.0.0.0", "127.0.0.1"} else host
                port = int((getattr(config, "server", None) and config.server.port) or 8090)
            if port in {80, 443}:
                return f"http://{host}"
            return f"http://{host}:{port}"
        except Exception:
            return "http://localhost"

    @staticmethod
    def _build_report_public_download_link(generation_id: int, expires_seconds: int = 86400) -> str:
        """生成报表公开下载链接（签名校验，默认24小时过期）。"""
        exp = int(datetime.now().timestamp()) + int(expires_seconds)
        raw = f"{generation_id}:{exp}".encode("utf-8")
        key = settings.SECRET_KEY.encode("utf-8")
        sig = hmac.new(key, raw, hashlib.sha256).hexdigest()
        base_url = NotifyTaskExecutor._guess_public_base_url()
        return f"{base_url}/api/v1/report/generation/public-download/{generation_id}?exp={exp}&sig={sig}"

    @staticmethod
    async def execute_report_send_task(task_id: int) -> Dict[str, Any]:
        if not await NotifyTaskExecutor._acquire_report_send_lock(task_id):
            return {"success": True, "message": "任务正在执行，跳过重复触发"}

        run_log = await NotifyTaskExecutor._create_run_log("report_send", task_id)
        task = await ReportSendTask.get_or_none(id=task_id).prefetch_related("report_config", "sender")
        if not task:
            await NotifyTaskExecutor._finish_run_log(run_log, TaskRunStatus.FAILED, error="任务不存在")
            await NotifyTaskExecutor._release_report_send_lock(task_id)
            return {"success": False, "message": "任务不存在"}

        try:
            sender = await task.sender
            if not sender or not sender.is_enabled:
                raise ValueError("发送人不存在或已禁用")

            report_config = await task.report_config
            if not report_config:
                raise ValueError("报表配置不存在")

            generation_name = f"{report_config.report_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            generation = await ReportGeneration.create(
                report_name=generation_name,
                report_config_id=report_config.id,
                generator="system-task",
                status="exporting",
            )

            # 执行生成并保留报表生成记录
            await ExcelExportService().export_report(generation.id)
            generation = await ReportGeneration.get(id=generation.id)
            if generation.status != "completed" or not generation.file_path:
                raise ValueError("报表生成失败")

            message = NotifyTaskExecutor._fill_report_message(
                template=task.message_template or "",
                report_name=report_config.report_name,
            )
            text_resp = WecomBotService.send_text(sender.channel_target, message)

            file_resp = None
            fallback_reason = None
            storage_meta = oss_service.extract_oss_meta(generation.execution_json)
            download_url = oss_service.resolve_download_url(storage_meta)
            if not download_url:
                download_url = NotifyTaskExecutor._build_report_public_download_link(generation_id=generation.id)

            local_file_path = None
            execution_json = generation.execution_json if isinstance(generation.execution_json, dict) else {}
            candidate_local_file = execution_json.get("local_file_path")
            if candidate_local_file and os.path.exists(candidate_local_file):
                local_file_path = candidate_local_file
            elif generation.file_path and os.path.exists(generation.file_path):
                local_file_path = generation.file_path

            try:
                if local_file_path:
                    file_resp = WecomBotService.send_file(sender.channel_target, local_file_path)
                else:
                    fallback_reason = "本地文件不存在或不可访问"
            except Exception as file_exc:
                fallback_reason = str(file_exc)

            if not file_resp:
                fallback_msg = (
                    f"{message}\n\n附件发送失败，已改为提供下载链接：\n{download_url}\n\n"
                    f"失败原因：{fallback_reason or '未知错误'}"
                )
                text_resp = WecomBotService.send_text(sender.channel_target, fallback_msg)

            task.last_run_time = datetime.now()
            await task.save()

            await NotifyTaskExecutor._create_alert_send_log(
                sender=sender,
                alert_text=message,
                send_status=1,
                response_text=str({"text": text_resp, "file": file_resp}),
            )
            await NotifyTaskExecutor._finish_run_log(
                run_log=run_log,
                status=TaskRunStatus.SUCCESS,
                output="执行成功" if file_resp else "执行成功（附件发送失败，已发送下载链接）",
                result_json={
                    "generation_id": generation.id,
                    "file_path": generation.file_path,
                    "local_file_path": local_file_path,
                    "download_url": download_url,
                    "file_sent": bool(file_resp),
                },
            )
            return {"success": True, "message": "执行成功", "generation_id": generation.id}
        except Exception as exc:
            task.last_run_time = datetime.now()
            await task.save()
            await NotifyTaskExecutor._finish_run_log(run_log, TaskRunStatus.FAILED, error=str(exc))
            sender = await AlertSender.get_or_none(id=task.sender_id) if task.sender_id else None
            if sender:
                await NotifyTaskExecutor._create_alert_send_log(
                    sender=sender,
                    alert_text=f"定时报表任务执行失败: {task.task_name}",
                    send_status=0,
                    response_text=str(exc),
                )
            return {"success": False, "message": str(exc)}
        finally:
            await NotifyTaskExecutor._release_report_send_lock(task_id)

    @staticmethod
    def _safe_str(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    @staticmethod
    def _build_rows_text(rows: List[dict], max_rows: int = 20) -> str:
        if not rows:
            return ""
        # 列名直接来自SQL查询结果（支持SQL别名，如中文列名）
        columns = list(rows[0].keys())
        lines = []
        display_rows = rows[:max_rows]
        for row in display_rows:
            parts = [f"{col}：{NotifyTaskExecutor._safe_str(row.get(col))}" for col in columns]
            lines.append("，".join(parts))
        if len(rows) > max_rows:
            lines.append("...")
        return "\n".join(lines)

    @staticmethod
    def _fill_message(template: str, rows: List[dict], total: int):
        msg = template or ""
        msg = msg.replace("{{total}}", str(total))
        msg = msg.replace("{total}", str(total))

        # 详情占位符：{{rows}} 或 {{detail_rows}}
        rows_text = NotifyTaskExecutor._build_rows_text(rows=rows, max_rows=NotifyTaskExecutor.MAX_MESSAGE_ROWS)
        msg = msg.replace("{{rows}}", rows_text)
        msg = msg.replace("{{detail_rows}}", rows_text)

        # 去掉其他遗留占位符，避免原样发出模板字符
        msg = re.sub(r"\{\{\s*[^}]+\s*\}\}", "", msg)
        return msg

    @staticmethod
    def _build_detail_excel(task_name: str, rows: List[dict]) -> Optional[str]:
        if not rows:
            return None
        file_dir = ExcelExportService()._get_file_dir()
        os.makedirs(file_dir, exist_ok=True)
        file_name = f"{task_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx".replace("/", "_")
        file_path = os.path.join(file_dir, file_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "detail"

        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        wb.save(file_path)
        return file_path

    @staticmethod
    async def execute_sql_alert_task(task_id: int) -> Dict[str, Any]:
        run_log = await NotifyTaskExecutor._create_run_log("sql_alert", task_id)
        task = await SqlAlertTask.get_or_none(id=task_id).prefetch_related("db_connection", "sender")
        if not task:
            await NotifyTaskExecutor._finish_run_log(run_log, TaskRunStatus.FAILED, error="任务不存在")
            return {"success": False, "message": "任务不存在"}

        try:
            sender = await task.sender
            if not sender or not sender.is_enabled:
                raise ValueError("发送人不存在或已禁用")

            db_conn = await task.db_connection
            if not db_conn:
                raise ValueError("数据库连接不存在")

            is_valid, validate_msg = ReportService.validate_sql(task.sql_statement)
            if not is_valid:
                raise ValueError(validate_msg)

            rows, total = await SQLExecutionService.execute_query(db_conn=db_conn, sql=task.sql_statement, offset=0, limit=2000)
            message = NotifyTaskExecutor._fill_message(
                template=task.message_template,
                rows=rows,
                total=total,
            )
            if not message:
                message = f"SQL预警结果：共 {total} 条"

            text_resp = WecomBotService.send_text(sender.channel_target, message)
            file_path = None
            file_resp = None
            if task.send_detail_excel and rows:
                file_path = NotifyTaskExecutor._build_detail_excel(task.task_name, rows)
                if file_path and os.path.exists(file_path):
                    file_resp = WecomBotService.send_file(sender.channel_target, file_path)

            task.last_run_time = datetime.now()
            await task.save()
            await NotifyTaskExecutor._create_alert_send_log(
                sender=sender,
                alert_text=message,
                send_status=1,
                response_text=str({"text": text_resp, "file": file_resp}),
            )
            await NotifyTaskExecutor._finish_run_log(
                run_log=run_log,
                status=TaskRunStatus.SUCCESS,
                output="执行成功",
                result_json={"total": total, "detail_file": file_path},
            )
            return {"success": True, "message": "执行成功", "total": total}
        except Exception as exc:
            task.last_run_time = datetime.now()
            await task.save()
            await NotifyTaskExecutor._finish_run_log(run_log, TaskRunStatus.FAILED, error=str(exc))
            sender = await AlertSender.get_or_none(id=task.sender_id) if task.sender_id else None
            if sender:
                await NotifyTaskExecutor._create_alert_send_log(
                    sender=sender,
                    alert_text=f"SQL预警任务执行失败: {task.task_name}",
                    send_status=0,
                    response_text=str(exc),
                )
            return {"success": False, "message": str(exc)}
