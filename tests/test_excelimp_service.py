"""
Tests for Excel Import Service
"""
import pytest
import sys
from pathlib import Path
from io import BytesIO
import openpyxl
from openpyxl import Workbook

# Add parent directory to path to import from app.services
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.excelimp_service import (
    _parse_sheet,
    _generate_field_names,
    _infer_field_types,
    _generate_table_name,
    _generate_create_table,
    _generate_insert_statements,
    _format_value,
    generate_sql
)


class TestParseSheet:
    """Tests for _parse_sheet function - Task 2.1"""
    
    def test_parse_sheet_basic(self):
        """Test parsing a basic Excel sheet with column names and data"""
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        
        # Add column names
        ws.append(["姓名", "年龄", "城市"])
        # Add data rows
        ws.append(["张三", 25, "北京"])
        ws.append(["李四", 30, "上海"])
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == ["姓名", "年龄", "城市"]
        assert len(data_rows) == 2
        assert data_rows[0] == ["张三", 25, "北京"]
        assert data_rows[1] == ["李四", 30, "上海"]
    
    def test_parse_sheet_with_none_values(self):
        """Test parsing sheet with None/empty values"""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", None, "NYC"])
        ws.append([None, 25, None])
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == ["Name", "Age", "City"]
        assert len(data_rows) == 2
        assert data_rows[0] == ["Alice", None, "NYC"]
        assert data_rows[1] == [None, 25, None]
    
    def test_parse_sheet_with_empty_rows(self):
        """Test that empty rows are skipped"""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["Col1", "Col2"])
        ws.append(["Data1", "Data2"])
        ws.append([None, None])  # Empty row
        ws.append(["", ""])  # Empty strings
        ws.append(["Data3", "Data4"])
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == ["Col1", "Col2"]
        assert len(data_rows) == 2  # Empty rows should be skipped
        assert data_rows[0] == ["Data1", "Data2"]
        assert data_rows[1] == ["Data3", "Data4"]
    
    def test_parse_sheet_with_none_column_names(self):
        """Test handling of None values in column names"""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["Col1", None, "Col3"])
        ws.append(["A", "B", "C"])
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == ["Col1", "column_2", "Col3"]
        assert len(data_rows) == 1
        assert data_rows[0] == ["A", "B", "C"]
    
    def test_parse_sheet_empty(self):
        """Test parsing an empty sheet"""
        wb = Workbook()
        ws = wb.active
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == []
        assert data_rows == []
    
    def test_parse_sheet_only_headers(self):
        """Test parsing sheet with only column names, no data"""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["Header1", "Header2", "Header3"])
        
        columns, data_rows = _parse_sheet(ws)
        
        assert columns == ["Header1", "Header2", "Header3"]
        assert data_rows == []


class TestGenerateFieldNames:
    """Tests for _generate_field_names function - Task 2.2"""
    
    def test_basic_field_names(self):
        """Test basic field name generation"""
        columns = ["Name", "Age", "City"]
        field_names = _generate_field_names(columns)
        
        assert field_names == ["Name", "Age", "City"]
    
    def test_field_names_with_spaces(self):
        """Test field names with spaces are converted to underscores"""
        columns = ["First Name", "Last Name", "Email Address"]
        field_names = _generate_field_names(columns)
        
        assert field_names == ["First_Name", "Last_Name", "Email_Address"]
    
    def test_field_names_with_special_chars(self):
        """Test field names with special characters"""
        columns = ["Name!", "Age@", "City#"]
        field_names = _generate_field_names(columns)
        
        # Special chars should be replaced with underscores
        assert all("_" in name or name.isalnum() for name in field_names)
    
    def test_duplicate_field_names(self):
        """Test handling of duplicate field names"""
        columns = ["Name", "Name", "Name"]
        field_names = _generate_field_names(columns)
        
        assert len(field_names) == 3
        assert len(set(field_names)) == 3  # All unique
        assert field_names[0] == "Name"
        assert field_names[1] == "Name_1"
        assert field_names[2] == "Name_2"


class TestInferFieldTypes:
    """Tests for _infer_field_types function - Task 2.3"""
    
    def test_infer_int_type(self):
        """Test inferring INT type"""
        data_rows = [[1, 2], [3, 4], [5, 6]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "INT"
        assert field_types[1] == "INT"
    
    def test_infer_bigint_type(self):
        """Test inferring BIGINT type for large integers"""
        data_rows = [[2147483648, 1], [2147483649, 2]]  # Larger than INT max
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "BIGINT"
    
    def test_infer_decimal_type(self):
        """Test inferring DECIMAL type for floats"""
        data_rows = [[1.5, 2.3], [3.7, 4.9]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DECIMAL(18,2)"
        assert field_types[1] == "DECIMAL(18,2)"
    
    def test_infer_varchar_type(self):
        """Test inferring VARCHAR type"""
        data_rows = [["Alice", "Bob"], ["Charlie", "David"]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert "VARCHAR" in field_types[0]
        assert "VARCHAR" in field_types[1]
    
    def test_infer_text_type(self):
        """Test inferring TEXT type for long strings"""
        long_string = "x" * 1500
        data_rows = [[long_string, "short"]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "TEXT"
    
    def test_infer_with_null_values(self):
        """Test type inference with NULL values"""
        data_rows = [[1, None], [2, None], [3, None]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "INT"
        assert "VARCHAR" in field_types[1]  # All NULL defaults to VARCHAR
    
    def test_infer_date_type(self):
        """Test inferring DATE type from date objects"""
        from datetime import date
        data_rows = [[date(2024, 1, 1), date(2024, 1, 2)], [date(2024, 1, 3), date(2024, 1, 4)]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DATE"
        assert field_types[1] == "DATE"
    
    def test_infer_datetime_type(self):
        """Test inferring DATETIME type from datetime objects"""
        from datetime import datetime
        data_rows = [
            [datetime(2024, 1, 1, 10, 30), datetime(2024, 1, 2, 14, 45)],
            [datetime(2024, 1, 3, 8, 15), datetime(2024, 1, 4, 16, 20)]
        ]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DATETIME"
        assert field_types[1] == "DATETIME"
    
    def test_infer_datetime_type_postgresql(self):
        """Test inferring TIMESTAMP type for PostgreSQL"""
        from datetime import datetime
        data_rows = [[datetime(2024, 1, 1, 10, 30)]]
        field_types = _infer_field_types(data_rows, "postgresql")
        
        assert field_types[0] == "TIMESTAMP"
    
    def test_infer_date_from_string(self):
        """Test inferring DATE type from date strings"""
        data_rows = [["2024-01-01", "2024-01-02"], ["2024-01-03", "2024-01-04"]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DATE"
        assert field_types[1] == "DATE"
    
    def test_infer_datetime_from_string(self):
        """Test inferring DATETIME type from datetime strings"""
        data_rows = [["2024-01-01 10:30:00", "2024-01-02 14:45:00"]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DATETIME"
        assert field_types[1] == "DATETIME"
    
    def test_infer_mixed_int_float(self):
        """Test mixed int and float values - should choose DECIMAL"""
        data_rows = [[1, 2.5], [3, 4.7], [5, 6]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        # First column is all int
        assert field_types[0] == "INT"
        # Second column has mixed int/float - should be DECIMAL
        assert field_types[1] == "DECIMAL(18,2)"
    
    def test_infer_mixed_int_string(self):
        """Test mixed int and string values - should choose VARCHAR"""
        data_rows = [[1, "text"], [2, "more"], [3, 4]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        # First column is all int
        assert field_types[0] == "INT"
        # Second column has mixed types - should be VARCHAR
        assert "VARCHAR" in field_types[1]
    
    def test_infer_mixed_date_datetime(self):
        """Test mixed date and datetime - should choose DATETIME (most permissive)"""
        from datetime import date, datetime
        data_rows = [[date(2024, 1, 1)], [datetime(2024, 1, 2, 10, 30)]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        assert field_types[0] == "DATETIME"
    
    def test_varchar_length_calculation(self):
        """Test VARCHAR length is calculated correctly with buffer"""
        data_rows = [["short"], ["a bit longer"], ["x" * 50]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        # Max length is 50, with 2x buffer should be 100
        assert field_types[0] == "VARCHAR(100)"
    
    def test_varchar_max_length_cap(self):
        """Test VARCHAR length is capped at 255"""
        data_rows = [["x" * 200]]
        field_types = _infer_field_types(data_rows, "mysql")
        
        # 200 * 2 = 400, but should be capped at 255
        assert field_types[0] == "VARCHAR(255)"


class TestFormatValue:
    """Tests for _format_value function"""
    
    def test_format_none(self):
        """Test formatting None values"""
        assert _format_value(None) == "NULL"
    
    def test_format_boolean(self):
        """Test formatting boolean values"""
        assert _format_value(True) == "TRUE"
        assert _format_value(False) == "FALSE"
    
    def test_format_numbers(self):
        """Test formatting numeric values"""
        assert _format_value(42) == "42"
        assert _format_value(3.14) == "3.14"
    
    def test_format_strings(self):
        """Test formatting string values"""
        assert _format_value("hello") == "'hello'"
    
    def test_format_strings_with_quotes(self):
        """Test formatting strings with single quotes (SQL escaping)"""
        assert _format_value("O'Brien") == "'O''Brien'"
    
    def test_format_date(self):
        """Test formatting date values"""
        from datetime import date
        d = date(2024, 1, 15)
        assert _format_value(d) == "'2024-01-15'"
    
    def test_format_datetime(self):
        """Test formatting datetime values"""
        from datetime import datetime
        dt = datetime(2024, 1, 15, 14, 30, 45)
        assert _format_value(dt) == "'2024-01-15 14:30:45'"


class TestGenerateTableName:
    """Tests for _generate_table_name function"""
    
    def test_table_name_format(self):
        """Test that table name has correct format"""
        table_name = _generate_table_name()
        
        assert table_name.startswith("tmp_")
        assert len(table_name) == 19  # tmp_YYYYMMDD_HHMMSS


class TestGenerateCreateTable:
    """Tests for _generate_create_table function"""
    
    def test_create_table_mysql(self):
        """Test CREATE TABLE for MySQL"""
        sql = _generate_create_table(
            "test_table",
            ["id", "name"],
            ["INT", "VARCHAR(50)"],
            "mysql"
        )
        
        assert "CREATE TABLE test_table" in sql
        assert "id INT" in sql
        assert "name VARCHAR(50)" in sql
        assert "ENGINE=InnoDB" in sql
        assert "CHARSET=utf8mb4" in sql
    
    def test_create_table_postgresql(self):
        """Test CREATE TABLE for PostgreSQL"""
        sql = _generate_create_table(
            "test_table",
            ["id", "name"],
            ["INT", "VARCHAR(50)"],
            "postgresql"
        )
        
        assert "CREATE TABLE test_table" in sql
        assert "id INT" in sql
        assert "name VARCHAR(50)" in sql
        assert "ENGINE" not in sql  # PostgreSQL doesn't use ENGINE


class TestGenerateInsertStatements:
    """Tests for _generate_insert_statements function"""
    
    def test_insert_single_batch(self):
        """Test INSERT statement generation for small dataset"""
        sql = _generate_insert_statements(
            "test_table",
            ["id", "name"],
            [[1, "Alice"], [2, "Bob"]],
            "mysql"
        )
        
        assert "INSERT INTO test_table" in sql
        assert "(id, name)" in sql
        assert "(1, 'Alice')" in sql
        assert "(2, 'Bob')" in sql
    
    def test_insert_multiple_batches(self):
        """Test INSERT statement generation with batching"""
        # Create 600 rows to test batching (batch_size=500)
        data_rows = [[i, f"Name{i}"] for i in range(600)]
        
        sql = _generate_insert_statements(
            "test_table",
            ["id", "name"],
            data_rows,
            "mysql",
            batch_size=500
        )
        
        # Should have 2 INSERT statements
        assert sql.count("INSERT INTO test_table") == 2
    
    def test_insert_with_null_values(self):
        """Test INSERT with NULL values"""
        sql = _generate_insert_statements(
            "test_table",
            ["id", "name"],
            [[1, None], [2, "Bob"]],
            "mysql"
        )
        
        assert "(1, NULL)" in sql
        assert "(2, 'Bob')" in sql


class TestGenerateSQL:
    """Integration tests for generate_sql function"""
    
    def test_generate_sql_complete(self):
        """Test complete SQL generation from Excel file"""
        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "年龄", "城市"])
        ws.append(["张三", 25, "北京"])
        ws.append(["李四", 30, "上海"])
        
        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Generate SQL
        sql = generate_sql(excel_bytes.read(), "test.xlsx", "mysql")
        
        # Verify SQL contains expected parts
        assert "CREATE TABLE tmp_" in sql
        assert "INSERT INTO tmp_" in sql
        assert "ENGINE=InnoDB" in sql
    
    def test_generate_sql_empty_file(self):
        """Test error handling for empty Excel file"""
        wb = Workbook()
        ws = wb.active
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        with pytest.raises(ValueError, match="没有找到列名"):
            generate_sql(excel_bytes.read(), "test.xlsx", "mysql")
    
    def test_generate_sql_no_data(self):
        """Test error handling for Excel with headers but no data"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Col1", "Col2"])
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        with pytest.raises(ValueError, match="没有找到数据行"):
            generate_sql(excel_bytes.read(), "test.xlsx", "mysql")
