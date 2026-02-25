"""
Minimal test script for Excel Import Service - Task 2.1
Tests the _parse_sheet function directly
"""
import sys
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook

# Import the module directly without going through app package
sys.path.insert(0, str(Path(__file__).parent.parent / "app" / "services"))
import excelimp_service


def test_parse_sheet_basic():
    """Test parsing a basic Excel sheet with column names and data"""
    wb = Workbook()
    ws = wb.active
    
    # Add column names
    ws.append(["姓名", "年龄", "城市"])
    # Add data rows
    ws.append(["张三", 25, "北京"])
    ws.append(["李四", 30, "上海"])
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == ["姓名", "年龄", "城市"], f"Expected ['姓名', '年龄', '城市'], got {columns}"
    assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"
    assert data_rows[0] == ["张三", 25, "北京"], f"Expected ['张三', 25, '北京'], got {data_rows[0]}"
    assert data_rows[1] == ["李四", 30, "上海"], f"Expected ['李四', 30, '上海'], got {data_rows[1]}"
    print("✓ test_parse_sheet_basic passed")


def test_parse_sheet_with_none_values():
    """Test parsing sheet with None/empty values"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", None, "NYC"])
    ws.append([None, 25, None])
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == ["Name", "Age", "City"]
    assert len(data_rows) == 2
    assert data_rows[0] == ["Alice", None, "NYC"]
    assert data_rows[1] == [None, 25, None]
    print("✓ test_parse_sheet_with_none_values passed")


def test_parse_sheet_with_empty_rows():
    """Test that empty rows are skipped"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Col1", "Col2"])
    ws.append(["Data1", "Data2"])
    ws.append([None, None])  # Empty row
    ws.append(["", ""])  # Empty strings
    ws.append(["Data3", "Data4"])
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == ["Col1", "Col2"]
    assert len(data_rows) == 2, f"Expected 2 data rows (empty rows skipped), got {len(data_rows)}"
    assert data_rows[0] == ["Data1", "Data2"]
    assert data_rows[1] == ["Data3", "Data4"]
    print("✓ test_parse_sheet_with_empty_rows passed")


def test_parse_sheet_with_none_column_names():
    """Test handling of None values in column names"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Col1", None, "Col3"])
    ws.append(["A", "B", "C"])
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == ["Col1", "column_2", "Col3"], f"Expected ['Col1', 'column_2', 'Col3'], got {columns}"
    assert len(data_rows) == 1
    assert data_rows[0] == ["A", "B", "C"]
    print("✓ test_parse_sheet_with_none_column_names passed")


def test_parse_sheet_empty():
    """Test parsing an empty sheet"""
    wb = Workbook()
    ws = wb.active
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == []
    assert data_rows == []
    print("✓ test_parse_sheet_empty passed")


def test_parse_sheet_only_headers():
    """Test parsing sheet with only column names, no data"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Header1", "Header2", "Header3"])
    
    columns, data_rows = excelimp_service._parse_sheet(ws)
    
    assert columns == ["Header1", "Header2", "Header3"]
    assert data_rows == []
    print("✓ test_parse_sheet_only_headers passed")


if __name__ == "__main__":
    print("Running Excel Import Service Tests - Task 2.1\n")
    print("Testing _parse_sheet function:\n")
    
    try:
        test_parse_sheet_basic()
        test_parse_sheet_with_none_values()
        test_parse_sheet_with_empty_rows()
        test_parse_sheet_with_none_column_names()
        test_parse_sheet_empty()
        test_parse_sheet_only_headers()
        
        print("\n" + "="*50)
        print("All tests passed! ✓")
        print("="*50)
        print("\nTask 2.1 Implementation Verified:")
        print("✓ Uses openpyxl to read Excel first sheet")
        print("✓ Extracts column names from first row")
        print("✓ Extracts data rows correctly")
        print("✓ Handles None/empty values properly")
        print("✓ Skips empty rows")
        print("✓ Handles None in column names")
        
    except AssertionError as e:
        print(f"\n✗ Test failed: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
